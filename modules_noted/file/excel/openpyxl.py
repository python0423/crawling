# -*- coding: utf-8 -*-
# @Author: admin
# @Date:   2018-03-19 12:27:01
# @Last Modified by:   admin
# @Last Modified time: 2018-03-20 12:03:31
# 此文档研究关于openpysl模块
# 这个模块用来处理excel文件
# 这个模块用来处理office07及以上的文件，03的没有试过
# xls和xlsx文件有很大的不同，后者是基于xml编写的
import openpyxl
def intro_openpyxl():
	# 命令行操作excel的方式
	# 介绍如何打开，关闭excel文件，以及一些打开关闭的属性
	#创建工作簿
	wb=Workbook()
	# 工作簿将至少创建一个工作表，可以通过active属性来获取第一个表
	ws=wb.active()

	# 可以使用create_sheet来创建新的工作表，参数是工作表的名字，后面是第几个表
	ws0=wb.create_sheet("my sheet",0)  #表示第一个位置的表为mysheet
	ws1=wb.create_sheet("my sheet two",1)
	# 可以使用title属性来改变表名
	ws0.title("my shit")
	# 可以通过sheetnames来查看所有表的名称
	print ws.sheetnames

	# 可以修改标题的背景颜色,颜色必须是rgb代码
	ws.sheet_properties.tabColor="1072BA"







if __name__ == '__main__':
	intro_openpyxl()
